import os
import zipfile
from PyPDF2 import PdfReader
from openpyxl import load_workbook
import pandas as pd

path_root = os.path.dirname(os.path.abspath(__file__))
path_tmp = os.path.join(path_root, 'tmp')
path_resources = os.path.join(path_root, 'resources')
path_resources_zip_file = os.path.join(path_root, 'resources', 'file.zip')
path_tmp_zip_file = os.path.join(path_root, 'tmp', 'file.zip')


def test_create_arch():
    os.chdir(path_resources)
    with zipfile.ZipFile('file.zip', 'w') as za:
        za.write('file1.txt')
        za.write('file2.xls')
        za.write('file3.xlsx')
        za.write('file4.pdf')
    if not os.path.isdir(path_tmp):
        os.mkdir(path_tmp)
    if os.path.exists(path_tmp_zip_file):
        os.remove(path_tmp_zip_file)
    os.rename(path_resources_zip_file, path_tmp_zip_file)


def test_txt():
    with zipfile.ZipFile(path_tmp_zip_file) as zip_f:
        with zip_f.open('file1.txt') as txt_f:
            assert txt_f.read().decode('utf-8') == 'Test string.'
            txt_file_size = zip_f.infolist()[0].file_size
            assert txt_file_size == 12


def test_xls():
    with zipfile.ZipFile(path_tmp_zip_file,'r') as zip_f:
        with zip_f.open("file2.xls") as xls_f:
            dt = pd.read_excel(xls_f).head()
            assert 'First Name' in dt
            xls_file_size = zip_f.infolist()[1].file_size
            assert xls_file_size ==  8704


def test_xlsx_abc():
    with zipfile.ZipFile(path_tmp_zip_file) as zip_f:
        with zip_f.open("file3.xlsx") as xlsx_data:
            workbook = load_workbook(xlsx_data)
            sheet = workbook.active
            sheet = sheet.cell(row=3, column=3).value
            assert sheet == 'Hashimoto'
            xlsx_file_size = zip_f.infolist()[2].file_size
            assert xlsx_file_size ==  7360


def test_pdf():
    with zipfile.ZipFile(path_tmp_zip_file) as zip_f:
        with zip_f.open('file4.pdf') as pdf_f:
            pdf_data = PdfReader(pdf_f)
            assert "Python Testing with pytest\n" in pdf_data.pages[1].extract_text()
            number_of_page = len(pdf_data.pages)
            assert number_of_page == 256
            pdf_file_size = zip_f.infolist()[3].file_size
            assert pdf_file_size == 3035139



